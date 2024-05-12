from openpyxl import load_workbook
from docx import Document 
from datetime import datetime
#passar os arquivos da planilha para o arquivo word
planilha_fornecedores = load_workbook("./fornecedores(1).xlsx")
pagina_fornecedores = planilha_fornecedores["Sheet 1"]

for linha in pagina_fornecedores.iter_rows(min_row = 2, values_only = True):
    nome_empresa, endereco, cidade, estado, cep, telefone, email, setor = linha  #umpacking

    arquivo_word = Document()
    arquivo_word.add_heading("Contrato de Prestação de serviço", 0)

    texto_contrato = f""""
      """
#salvar o arquivo word em uma pasta especifica(Contratos)
#repetir para todas as linhas da planilha